"""Unit tests for the mechanical Excel processor."""
import datetime
import io
from unittest.mock import patch

import numpy as np
import pandas as pd
from openpyxl import Workbook

from processing.excel import (
    allowed_file,
    build_nested_columns,
    clean_header_values,
    convert_to_serializable,
    is_blank_header_value,
    is_empty_or_irrelevant,
    parse_excel_with_decisions,
)


def test_allowed_file_accepted():
    assert allowed_file("report.xlsx") is True
    assert allowed_file("data.xls") is True
    assert allowed_file("macro.xlsm") is True


def test_allowed_file_rejected():
    assert allowed_file("note.txt") is False
    assert allowed_file("noextension") is False
    assert allowed_file("") is False


def test_convert_to_serializable_datetime():
    dt = datetime.datetime(2024, 6, 1, 12, 30, 0)
    assert convert_to_serializable(dt) == dt.isoformat()
    d = datetime.date(2024, 6, 1)
    assert convert_to_serializable(d) == d.isoformat()


def test_convert_to_serializable_numpy():
    assert convert_to_serializable(np.int64(42)) == 42
    assert convert_to_serializable(np.float64(3.5)) == 3.5
    assert convert_to_serializable(np.bool_(True)) is True
    arr = np.array([1, 2])
    assert convert_to_serializable(arr) == [1, 2]


def test_convert_to_serializable_nested():
    payload = {
        "nums": np.array([1.0]),
        "nested": [np.int32(7), None],
    }
    out = convert_to_serializable(payload)
    assert out == {"nums": [1.0], "nested": [7, None]}


def test_build_nested_columns_merges_forward_fill():
    df = pd.DataFrame(
        [
            ["A", "A", "B"],
            ["x", "y", "z"],
        ]
    )
    cols = build_nested_columns(df, header_rows=2)
    assert cols[0] == ["A", "x"]
    assert cols[1] == ["A", "y"]
    assert cols[2] == ["B", "z"]


def test_build_nested_columns_treats_unnamed_and_untitled_as_empty():
    df = pd.DataFrame(
        [
            ["Target", "Untitled: 1", "Source", "Unnamed: 3"],
            ["table", "column", "table", "column"],
        ]
    )
    cols = build_nested_columns(df, header_rows=2)
    assert cols == [
        ["Target", "table"],
        ["Target", "column"],
        ["Source", "table"],
        ["Source", "column"],
    ]


def test_build_nested_columns_empty():
    assert build_nested_columns(pd.DataFrame(), header_rows=1) == []


def test_clean_header_values_removes_generated_on_tail():
    headers = ["Table", "Unnamed: 1", "Generated on 2026-07-20", "extra"]
    assert clean_header_values(headers) == ["Table"]


def test_is_blank_header_value_accepts_untitled_and_unnamed():
    assert is_blank_header_value("Untitled: 4") is True
    assert is_blank_header_value("Unnamed: 4") is True
    assert is_blank_header_value("Real column") is False


def test_is_empty_or_irrelevant_empty_list():
    ok, reason = is_empty_or_irrelevant([])
    assert ok is True
    assert "empty" in reason.lower()


def test_is_empty_or_irrelevant_whitespace_only():
    ok, reason = is_empty_or_irrelevant([[" ", ""], [None, None]])
    assert ok is True
    assert "whitespace" in reason.lower()


def test_is_empty_or_irrelevant_has_data():
    ok, reason = is_empty_or_irrelevant([[None, "Name"], [1, 2]])
    assert ok is False
    assert reason == ""


def test_parser_reads_each_sheet_once(sample_excel_bytes):
    with (
        patch("processing.excel.get_header_decision", return_value=(0, 1, False)),
        patch("processing.excel.pd.read_excel", wraps=pd.read_excel) as read_excel,
    ):
        sheets = parse_excel_with_decisions(sample_excel_bytes)

    assert read_excel.call_count == 1
    assert sheets[0]["columns"] == ["Name", "Age"]
    assert sheets[0]["header"] == {
        "start_row": 0,
        "row_count": 1,
        "nested": False,
    }
    assert sheets[0]["data_rows"] == [["Alice", 30], ["Bob", 25]]


def test_parser_uses_automatic_header_decision_for_untitled_rows():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Automatic"
    sheet.append(["Untitled: 0", None, "Unnamed: 2"])
    sheet.append(["table", "column", "description"])
    sheet.append(["T_CLIENT", "CLIENT_ID", "Client identifier"])
    output = io.BytesIO()
    workbook.save(output)

    with patch(
        "processing.excel.get_header_decision",
        return_value=(1, 1, False),
    ) as get_decision:
        sheets = parse_excel_with_decisions(output.getvalue())

    get_decision.assert_called_once()
    assert sheets[0]["header"] == {
        "start_row": 1,
        "row_count": 1,
        "nested": False,
    }


def test_parser_expands_only_real_merged_data_cells():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "S2T"
    sheet.append(["Target", "Target", "Transformation"])
    sheet.append(["Table", "Field", "Rule"])
    sheet.append(["T_TEST", "FIELD_A", "shared rule"])
    sheet.append(["T_TEST", "FIELD_B", None])
    sheet.append(["T_TEST", "FIELD_C", None])
    sheet.append(["T_TEST", "FIELD_D", None])
    sheet.merge_cells("C3:C5")
    output = io.BytesIO()
    workbook.save(output)

    with patch(
        "processing.excel.get_header_decision",
        return_value=(0, 2, True),
    ):
        sheets = parse_excel_with_decisions(output.getvalue())

    assert sheets[0]["data_rows"] == [
        ["T_TEST", "FIELD_A", "shared rule"],
        ["T_TEST", "FIELD_B", "shared rule"],
        ["T_TEST", "FIELD_C", "shared rule"],
        ["T_TEST", "FIELD_D", None],
    ]
    assert sheets[0]["data_row_numbers"] == [0, 1, 2, 3]


def test_parser_excludes_hidden_rows_by_default_and_preserves_row_numbers():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "S2T"
    sheet.append(["Target", "Target", "Transformation"])
    sheet.append(["Table", "Field", "Rule"])
    sheet.append(["T_TEST", "FIELD_A", "shared rule"])
    sheet.append(["T_TEST", "FIELD_B", None])
    sheet.append(["T_TEST", "FIELD_C", None])
    sheet.append(["T_TEST", "FIELD_D", None])
    sheet.merge_cells("C3:C5")
    sheet.row_dimensions[3].hidden = True
    sheet.row_dimensions[5].hidden = True
    output = io.BytesIO()
    workbook.save(output)
    with patch(
        "processing.excel.get_header_decision",
        return_value=(0, 2, True),
    ):
        visible_sheets = parse_excel_with_decisions(output.getvalue())
        all_sheets = parse_excel_with_decisions(
            output.getvalue(),
            include_hidden_rows=True,
        )

    assert visible_sheets[0]["data_rows"] == [
        ["T_TEST", "FIELD_B", "shared rule"],
        ["T_TEST", "FIELD_D", None],
    ]
    assert visible_sheets[0]["data_row_numbers"] == [1, 3]
    assert all_sheets[0]["data_rows"] == [
        ["T_TEST", "FIELD_A", "shared rule"],
        ["T_TEST", "FIELD_B", "shared rule"],
        ["T_TEST", "FIELD_C", "shared rule"],
        ["T_TEST", "FIELD_D", None],
    ]
    assert all_sheets[0]["data_row_numbers"] == [0, 1, 2, 3]
